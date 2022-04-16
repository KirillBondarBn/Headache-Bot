import openpyxl
from openpyxl.styles import Fill, PatternFill, Border, Side, Alignment
from datetime import datetime
from math import ceil
import os

def get_month_name(month):
    if month=='1':
        return 'Январь'
    elif month=='2':
        return 'Февраль'
    elif month == '3':
        return 'Март'
    elif month == '4':
        return 'Апрель'
    elif month == '5':
        return 'Май'
    elif month == '6':
        return 'Июнь'
    elif month == '7':
        return 'Июль'
    elif month == '8':
        return 'Август'
    elif month == '9':
        return 'Сентябрь'
    elif month == '10':
        return 'Октябрь'
    elif month == '11':
        return 'Ноябрь'
    elif month == '12':
        return 'Декабрь'
def get_day_xls_char(day):
    if day==1:
        return 'B'
    elif day==2:
        return 'C'
    elif day==3:
        return 'D'
    elif day==4:
        return 'E'
    elif day==5:
        return 'F'
    elif day==6:
        return 'G'
    elif day==7:
        return 'H'

def delete_report(name):
    os.remove(name)

def create_report(list_of_notes, user_id):
    # Создал xls файл
    book = openpyxl.Workbook()
    # Удалил изначально сгенерировавшийся sheet
    book.remove(book.active)

    years = []
    for i in list_of_notes:
        if (str(i[1])[0:4]) not in years:
            years.append(str(i[1])[0:4])

    for i in range(len(years)):
        for g in range(12):
            month = get_month_name(str(g + 1))
            sheet = book.create_sheet(month + ' ' + years[i])

            # style
            weekends_name_background_color = PatternFill('solid', fgColor='BBDBC4')
            weekday_name_background_color = PatternFill('solid', fgColor='BBD3DB')
            # weekends_background_color = PatternFill('solid', fgColor='##a3a3a3')

            sheet.column_dimensions['A'].width = 4
            sheet.column_dimensions['B'].width = 37
            sheet.column_dimensions['C'].width = 37
            sheet.column_dimensions['D'].width = 37
            sheet.column_dimensions['E'].width = 37
            sheet.column_dimensions['F'].width = 37
            sheet.column_dimensions['G'].width = 37
            sheet.column_dimensions['H'].width = 37

            alignment = Alignment(horizontal='general', vertical='center')

            side = Side(style='thin', color='000000')
            for a in 'BCDEFGH':
                for n in '234567':
                    sheet[a+n].border = Border(bottom=side, left=side, top=side, right=side)
                    sheet[a+n].alignment = alignment

            sheet['B2'].fill = weekday_name_background_color
            sheet['C2'].fill = weekday_name_background_color
            sheet['D2'].fill = weekday_name_background_color
            sheet['E2'].fill = weekday_name_background_color
            sheet['F2'].fill = weekday_name_background_color
            sheet['G2'].fill = weekends_name_background_color
            sheet['H2'].fill = weekends_name_background_color

            sheet.row_dimensions[3].height = 40
            sheet.row_dimensions[4].height = 40
            sheet.row_dimensions[5].height = 40
            sheet.row_dimensions[6].height = 40
            sheet.row_dimensions[7].height = 40

            sheet['B2'] = 'Понедельник'
            sheet['C2'] = 'Вторник'
            sheet['D2'] = 'Среда'
            sheet['E2'] = 'Четверг'
            sheet['F2'] = 'Пятница'
            sheet['G2'] = 'Суббота'
            sheet['H2'] = 'Воскресенье'

            sheet['B10'] = 'Среднее значение боли:'

            if g < 9:
                first_day_of_month = years[i][2:]+'/'+'0'+str(g+1)+'/'+'01'
            else:
                first_day_of_month = years[i][2:]+'/'+str(g+1) + '/' + '01'

            month_start = datetime.strptime(first_day_of_month, '%y/%m/%d')
            first_week_duration = 7 - month_start.weekday()
            ok = False
            cnt = summ = 0
            for j in list_of_notes:
                if ((str(j[1])[5:7]) == '0'+str(g+1) or (str(j[1])[5:7]) == str(g+1)) and (str(j[1])[0:4]) == years[i]:
                    line = str(j[1][2:16]).replace('-', '/')
                    date = datetime.strptime(line, '%y/%m/%d %H:%M')
                    week = ceil((int(line[6:8]) - first_week_duration) / 7) + 1
                    num_of_cell = get_day_xls_char(date.isoweekday()) + str(week+2)

                    if str(sheet[num_of_cell].value) == 'None':
                        current_cell_value = ''
                    else:
                        current_cell_value = str(sheet[num_of_cell].value)

                    sheet[num_of_cell].value = current_cell_value + str(j[2]) + ' ' + j[3]+' '+str(j[1])[11:16]+' | '
                    cnt += 1
                    summ += int(j[2])
                    ok = True
            if ok:
                sheet['C10'].value = str(summ/cnt)[:4]
                book.save(str(user_id)+'_report.xls')
            else:
                book.remove(sheet)